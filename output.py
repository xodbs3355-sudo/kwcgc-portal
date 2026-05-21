"""
검토 결과 엑셀 출력 + 첨부 파일 PDF 병합
"""
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image
from pypdf import PdfWriter, PdfReader


STATUS_COLORS = {
    "OK":   "C6EFCE",  # 연두
    "NG":   "FFC7CE",  # 연빨강
    "WARN": "FFEB9C",  # 연노랑
    "SKIP": "D9D9D9",  # 회색
}


def results_to_excel(all_results: dict[str, list[dict]]) -> bytes:
    """
    all_results: {"시트명": [{"항목":..., "결과":..., "추출값":..., "비고":...}, ...]}
    반환: xlsx bytes
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, rows in all_results.items():
        ws = wb.create_sheet(title=sheet_name[:31])

        # 헤더
        headers = ["항목", "결과", "추출값", "비고"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 데이터
        for row_idx, item in enumerate(rows, 2):
            status = item.get("결과", "SKIP")
            fill_color = STATUS_COLORS.get(status, "FFFFFF")
            fill = PatternFill("solid", fgColor=fill_color)

            for col_idx, key in enumerate(["항목", "결과", "추출값", "비고"], 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=item.get(key, ""))
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="center")

        # 열 너비
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 30

    # 요약 시트
    _add_summary_sheet(wb, all_results)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _add_summary_sheet(wb: openpyxl.Workbook, all_results: dict[str, list[dict]]):
    ws = wb.create_sheet(title="검토요약", index=0)

    ws.cell(1, 1, "준공서류 검토 결과 요약").font = Font(bold=True, size=14)
    ws.cell(2, 1, f"검토일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    ws.cell(4, 1, "시트명").font = Font(bold=True)
    ws.cell(4, 2, "OK").font = Font(bold=True)
    ws.cell(4, 3, "NG").font = Font(bold=True)
    ws.cell(4, 4, "WARN").font = Font(bold=True)
    ws.cell(4, 5, "최종").font = Font(bold=True)

    for row_idx, (sheet_name, rows) in enumerate(all_results.items(), 5):
        counts = {"OK": 0, "NG": 0, "WARN": 0, "SKIP": 0}
        for item in rows:
            counts[item.get("결과", "SKIP")] = counts.get(item.get("결과", "SKIP"), 0) + 1

        final = "합격" if counts["NG"] == 0 else "불합격"
        final_color = "C6EFCE" if final == "합격" else "FFC7CE"

        ws.cell(row_idx, 1, sheet_name)
        ws.cell(row_idx, 2, counts["OK"])
        ws.cell(row_idx, 3, counts["NG"])
        ws.cell(row_idx, 4, counts["WARN"])
        cell = ws.cell(row_idx, 5, final)
        cell.fill = PatternFill("solid", fgColor=final_color)
        cell.font = Font(bold=True)


# ── 첨부 파일 PDF 병합 ────────────────────────────────────────────
def _image_to_pdf_bytes(img_bytes: bytes) -> bytes:
    """이미지 bytes → PDF bytes 변환"""
    img = Image.open(io.BytesIO(img_bytes))
    if img.mode in ("RGBA", "P"):
        img = img.convert("RGB")
    buf = io.BytesIO()
    img.save(buf, format="PDF")
    return buf.getvalue()


def merge_attachments_to_pdf(uploaded_files: dict, documents: list) -> bytes:
    """
    uploaded_files: {doc_id: [(filename, bytes), ...]}
    documents: DOCUMENTS list (서류 순서 보장용)
    반환: 통합 PDF bytes
    """
    writer = PdfWriter()

    for doc in documents:
        did = doc["id"]
        files = uploaded_files.get(did, [])
        if not files:
            continue

        for filename, file_bytes in files:
            ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
            try:
                if ext == "pdf":
                    pdf_bytes = file_bytes
                elif ext in ("jpg", "jpeg", "png"):
                    pdf_bytes = _image_to_pdf_bytes(file_bytes)
                else:
                    continue

                reader = PdfReader(io.BytesIO(pdf_bytes))
                for page in reader.pages:
                    writer.add_page(page)
            except Exception:
                continue

    buf = io.BytesIO()
    writer.write(buf)
    return buf.getvalue()
