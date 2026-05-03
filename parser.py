"""
엑셀 파일 파싱 및 시트 이미지 추출 (.xlsx / .xls 모두 지원)
"""
import io

import openpyxl
from PIL import Image


# ── .xls용 워크시트 래퍼 ──────────────────────────────────────────
class XlsSheet:
    """xlrd 시트를 openpyxl 인터페이스처럼 감싸는 래퍼"""

    def __init__(self, xlrd_sheet):
        self._sheet = xlrd_sheet
        self._images = []  # .xls는 이미지 감지 불가

    def iter_rows(self, max_row=None, values_only=True):
        nrows = self._sheet.nrows
        if max_row:
            nrows = min(nrows, max_row)
        for r in range(nrows):
            row = [self._sheet.cell_value(r, c) for c in range(self._sheet.ncols)]
            yield row

    @property
    def max_row(self):
        return self._sheet.nrows


class XlsWorkbook:
    """xlrd 워크북을 openpyxl 인터페이스처럼 감싸는 래퍼"""

    def __init__(self, xlrd_wb):
        self._wb = xlrd_wb
        self.sheetnames = xlrd_wb.sheet_names()

    def __getitem__(self, name):
        return XlsSheet(self._wb.sheet_by_name(name))


# ── 워크북 로드 ───────────────────────────────────────────────────
def load_workbook(file_bytes: bytes, filename: str = '') -> openpyxl.Workbook:
    if filename.lower().endswith('.xls'):
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes, encoding_override='utf-8')
        return XlsWorkbook(wb)
    return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)


def get_sheet_names(wb) -> list[str]:
    return wb.sheetnames


def extract_images_from_sheet(ws) -> list[Image.Image]:
    images = []
    if not hasattr(ws, '_images'):
        return images
    for img in ws._images:
        try:
            pil_img = Image.open(io.BytesIO(img._data()))
            images.append(pil_img)
        except Exception:
            pass
    return images


def sheet_to_image(ws, scale: float = 1.5) -> Image.Image:
    from PIL import ImageDraw, ImageFont

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return Image.new("RGB", (400, 100), "white")

    col_count = max(len(r) for r in rows)
    cell_w, cell_h = 100, 22

    img_w = col_count * cell_w + 20
    img_h = len(rows) * cell_h + 20

    img = Image.new("RGB", (min(img_w, 2000), min(img_h, 3000)), "white")
    draw = ImageDraw.Draw(img)

    try:
        font = ImageFont.truetype("malgun.ttf", 11)
    except Exception:
        font = ImageFont.load_default()

    for row_idx, row in enumerate(rows):
        y = 10 + row_idx * cell_h
        for col_idx, val in enumerate(row):
            x = 10 + col_idx * cell_w
            val_str = str(val) if val is not None else ""
            draw.rectangle([x, y, x + cell_w - 1, y + cell_h - 1], outline="#cccccc")
            draw.text((x + 2, y + 3), val_str[:12], fill="black", font=font)

    return img


def extract_sheet_content(ws) -> dict:
    data = {}
    if hasattr(ws, 'iter_rows'):
        for row in ws.iter_rows(values_only=True):
            pass
    return data
